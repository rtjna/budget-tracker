import csv
import io
from datetime import datetime, time

from openpyxl import load_workbook

XLSX_MAGIC = b"PK\x03\x04"


class XlsxError(Exception):
    """The upload looked like an xlsx (zip magic) but can't be read as one."""


def is_xlsx(data: bytes) -> bool:
    return data[:4] == XLSX_MAGIC


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.strftime("%d/%m/%Y")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def xlsx_to_csv_text(data: bytes) -> str:
    """Flatten the first worksheet to CSV text so xlsx uploads flow through
    the same importer detection and parsing as CSV uploads. Raises XlsxError
    with a friendly message when the file isn't a readable workbook."""
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if not workbook.worksheets:
            raise XlsxError("The Excel workbook contains no worksheets")
        sheet = workbook.worksheets[0]
        out = io.StringIO()
        writer = csv.writer(out)
        for row in sheet.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            writer.writerow([_cell_to_str(cell) for cell in row])
        workbook.close()
    except XlsxError:
        raise
    except Exception as e:
        # openpyxl/zipfile raise a zoo of exceptions on truncated or corrupt
        # archives; collapse them into one clear message.
        raise XlsxError("Could not read this file as an Excel workbook (corrupt or not an xlsx?)") from e
    return out.getvalue()
